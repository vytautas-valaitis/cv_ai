import openpyxl, json
import doc_parser.constant as constant

deviationY = 10

def getAllWordsFromBlocks(startBlock, endBlock, data):
  allWords = []
  for blockInfo in data["pages"][0]["blocks"][startBlock:endBlock]:
    for paragraph in range(len(blockInfo["paragraphs"])):
      index = 0
      while index < len(blockInfo["paragraphs"][paragraph]["words"]):
        # susirašom visus žodžius į masyvą, kad būtų lengviau atlikti visus tikrinimus
        wordInfo = blockInfo["paragraphs"][paragraph]["words"][index]
        matchPage = constant.PATTERN_PAGE.search(wordInfo["text"])
        if matchPage is not None:
          # Jei surandame Page x of x, praleidžiam
          index += 3
        else:
          allWords.append(blockInfo["paragraphs"][paragraph]["words"][index])
        index += 1

  return allWords


def analyseExperienceWords(allWords, skillsJSON, cvIndex, excelSheet, rowInExcel, multiplePositionsInfo):
  experienceCount = 0
  fullText = ""
  fourLinesText = [allWords[0]["text"] + " ", "", "", ""]  # Išsaugom pirmą žodį, jog index galėtų būti iškart 1
  lineIndex = 0
  wordInlinePositionY = allWords[0]["positions"][0]["y"]
  isFirstDate = True  # Naudojama, jog praleistume pirmą rastą darbovietės datą
  isNewSkillRecordEnd = False

  index = 1
  while index < len(allWords):
    currentWord = allWords[index]["text"]
    matchPage = constant.PATTERN_PAGE.search(currentWord)
    if matchPage is not None:
      # Suradus tekstą 'page x of x' jį praleidžiam
      index += 4
      if index >= len(allWords):
        # Suveiks, jei page x of x bus paskutinis tekstas
        index = len(allWords) - 1
        # Lieka bug'as, jei pats paskutinis tekstas yra 'page x of x', tai įtrauks paskutinį sk į saugomą tekstą

    yPos = allWords[index]["positions"][0]["y"]
    if wordInlinePositionY + deviationY > yPos and wordInlinePositionY - deviationY < yPos:
      # Suveikia, jei žodis vis dar toje pačioje eilutėje
      fourLinesText[lineIndex] = addTextAndCheckSpaces(
        fourLinesText[lineIndex], currentWord, allWords[index - 1]["text"]
      )  # Pridedam žodį prie atitinkamos eilutės
      if lineIndex > 1 and index + 1 != len(allWords):
        # Jei jau yra užpildytos 2 eilutės, ieškome ar pasibaigė darbovietės aprašas
        matchYear = constant.PATTERN_YEAR.search(allWords[index - 1]["text"])
        matchPresent = constant.PATTERN_PRESENT_POSITION.search(allWords[index - 1]["text"])
        matchParentheses = constant.PATTERN_PARENTHESES.search(currentWord)
        matchNumber = constant.PATTERN_NUMBER.search(allWords[index + 1]["text"])
        if (
          matchParentheses is not None
          and matchNumber is not None
          and (matchYear is not None or matchPresent is not None)
        ):
          # Jei surandam metus(arba dabartinę poziciją), ( ir skaičių, žinom, kad šitos trys eilutės tinka
          if not isFirstDate:
            isNewSkillRecordEnd = True
          if isFirstDate:
            isFirstDate = False
      if index + 1 == len(allWords):
        # Pasiekus paskutinį žodį išsaugom į naują darbovietę
        if isNewSkillRecordEnd:
          fullText += fourLinesText[0]
          fourLinesText[0] = ""
          fullText = removeSpacesAndSymbols(fullText)
          experienceCount = addSkillsRecord(experienceCount, fullText, skillsJSON, cvIndex, excelSheet, rowInExcel)
          fullText = ""
        fullText += fourLinesText[0] + fourLinesText[1] + fourLinesText[2] + fourLinesText[3]
        fullText = removeSpacesAndSymbols(fullText)
        experienceCount = addSkillsRecord(experienceCount, fullText, skillsJSON, cvIndex, excelSheet, rowInExcel)
    else:
      isFirstFromMultiplePositions = True
      wordInlinePositionY = yPos
      index -= 1
      if lineIndex > 1:
        if isNewSkillRecordEnd:
          if experienceCount != len(multiplePositionsInfo):
            if not (not multiplePositionsInfo[experienceCount] and multiplePositionsInfo[experienceCount + 1]):
              # Jei sekantis experience nėra pirmas iš multiple positions įrašom papildomą eilutę
              # Kadangi pirmą įrašą sudaro sudaro 4 eilutės ir tuomet įrašyti nereikia
              isFirstFromMultiplePositions = False
              fullText = addLineTextAndRearrangeList(fullText, fourLinesText, lineIndex)
              if multiplePositionsInfo[experienceCount] and multiplePositionsInfo[experienceCount + 1]:
                # Jei tai yra antras ir daugiau experience iš multiple positions įrašom eilutę
                # Kadangi jį sudaro 2 eilutės
                fullText = addLineTextAndRearrangeList(fullText, fourLinesText, lineIndex)
          fullText = removeSpacesAndSymbols(fullText)
          experienceCount = addSkillsRecord(experienceCount, fullText, skillsJSON, cvIndex, excelSheet, rowInExcel)
          fullText = ""
          if isFirstFromMultiplePositions:
            # Pridedam teksto eilutę jau prie naujo experience
            fullText = addLineTextAndRearrangeList(fullText, fourLinesText, lineIndex)
          isNewSkillRecordEnd = False
        else:
          fullText = addLineTextAndRearrangeList(fullText, fourLinesText, lineIndex)
      if lineIndex != 3:
        lineIndex += 1

    index += 1


def addLineTextAndRearrangeList(fullText, textList, lineIndex):
  fullText += textList[0]
  textList[0] = textList[1]
  textList[1] = textList[2]
  textList[2] = textList[3]
  textList[3] = ""

  return fullText


def removeSpacesAndSymbols(text):
  for symbol in range(len(text) - 1, -1, -1):
    if not text[symbol].isalnum():
      # Jei paskutinis simbolis ne skaičius ir ne raidė - pašalinam
      text = text[:-1]
    else:
      break
  return text


def addSkillsRecord(experienceCount, fullText, jsonAI, cvIndex, excelSheet, rowInExcel):
  jsonAI["CVs"][cvIndex]["jobs"].append({})
  jsonAI["CVs"][cvIndex]["jobs"][experienceCount]["skillsFromCV"] = fullText
  jsonAI["CVs"][cvIndex]["jobs"][experienceCount]["skillsFromExcel"] = excelSheet.cell(
    row=rowInExcel, column=23 + experienceCount * 5
  ).value  # Skills'ai prasideda 23 stulpely ir kiekvieno darbo einam +5 stulpelyje
  return experienceCount + 1


def addTextAndCheckSpaces(allText, textToAdd, wordBefore):
  # Pašalinam tarpus prieš ir po atitinkamų simbolių ir pridedam tekstą prie string'o
  matchYear = constant.PATTERN_YEAR.search(wordBefore)
  matchNoLeftSpace = constant.PATTERN_SYMBOL_NO_LEFT_SPACE.search(textToAdd)
  matchNoRightSpace = constant.PATTERN_SYMBOL_NO_RIGHT_SPACE.search(textToAdd)
  matchNoSpaces = constant.PATTERN_SYMBOL_NO_SPACES.search(textToAdd)
  if matchNoLeftSpace is not None or matchNoSpaces is not None or (textToAdd == "-" and matchYear is None):
    allText = allText[:-1]

  allText += textToAdd + " "

  if (matchNoSpaces is not None or matchNoRightSpace is not None) or (textToAdd == "-" and matchYear is None):
    allText = allText[:-1]

  return allText


def getExcelSheetAndRow(cvFileName):
  wb = openpyxl.load_workbook("/opt/cv_data/CV duomenys 2021.xlsx")
  sheet = wb.active

  for row in range(sheet.max_row - 1):
    if cvFileName == sheet.cell(row=row + 1, column=1).value:
      rowInExcel = row + 1

  return sheet, rowInExcel


def getSkillsJSON(tmp_dir, cvIndex):
  if cvIndex == 0:
    skillsJSON = {}
    skillsJSON["CVs"] = []
  else:
    with open(tmp_dir + "/allSkills.json", "r") as dataFile:
      skillsJSON = json.load(dataFile)

  return skillsJSON


def fillSkills(tmp_dir, cvFileName, cvIndex, multiplePositionsInfo):
  skillsJSON = getSkillsJSON(tmp_dir, cvIndex)
  excelSheet, rowInExcel = getExcelSheetAndRow(cvFileName)

  skillsJSON["CVs"].append({})
  skillsJSON["CVs"][cvIndex]["jobs"] = []
  skillsJSON["CVs"][cvIndex]["FileName"] = cvFileName

  with open(tmp_dir + "/rightModified.json", "r") as dataFile:
    data = json.load(dataFile)

  for block in range(len(data["pages"][0]["blocks"])):
    # Surandam experiene pradžios ir pabaigos blokus
    matchEducation = constant.PATTERN_EDUCATION.search(data["pages"][0]["blocks"][block]["text"])
    matchExperience = constant.PATTERN_EXPERIENCE.search(data["pages"][0]["blocks"][block]["text"])
    if matchEducation is not None:
      experienceEndBlock = block
    if matchExperience is not None:
      experienceStartBlock = block + 1

  if "experienceEndBlock" not in locals():
    # Suveiks, jei CV nebus education dalies. Tuomet imam visus blokus iki paskutinio
    experienceEndBlock = len(data["pages"][0]["blocks"]) - 1
  if "experienceStartBlock" in locals() and "rowInExcel" in locals():
    allWords = getAllWordsFromBlocks(experienceStartBlock, experienceEndBlock, data)
    analyseExperienceWords(allWords, skillsJSON, cvIndex, excelSheet, rowInExcel, multiplePositionsInfo)
    # print(jsonAI)
    with open(tmp_dir + "/allSkills.json", "w") as dataFile:
      json.dump(skillsJSON, dataFile)
  else:
    # Suveiks, jei nebus experience dalies arba CV nebus excel faile (dalis pav. būna suvesti su klaidomis)
    print("PRALEISTAS: ", cvFileName)

